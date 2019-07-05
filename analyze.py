#!python3
import os
import re
import sys
import graphs
import openpyxl
from collections import Counter, OrderedDict
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle

#(Date)(, )(Hours:Minutes)(Seconds)(AM/PM)(:/ -)( )(Sender)(: )(Message)
message_format = re.compile(r'^(\[?)([0-9]{1,2}/[0-9]{2}/[0-9]{2,4})( |, )([0-9]{1,2}:[0-9]{2})(:[0-9]{2})?(]|( [A-Z]{2})?( -|:))( )(.*?)(: )(.*)')
# ([)(Day/Month/Year)( )(Hours:Minutes)(:Seconds)(] )(Sender)(: )(Message)
# (\[)([0-9]{1,2}/[0-9]{2}/[0-9]{2,4})( )([0-9]{1,2}:[0-9]{2})(:[0-9]{2})(])( )(.*?)(: )(.*)
time_split = re.compile(r'([0-9]{1,2}):([0-9]{2})')

#To get rid of file extension when making graphs
file_split = re.compile(r'(.*)(.[a-zA-Z0-9]{3,4})')

"""
If value exists in the dictionary, increment count. Else add a new entry.
"""
def add_to_dictionary(dictionary, value):
    if value in dictionary:
        dictionary[value] += 1
    else:
        dictionary[value] = 1
    return dictionary

"""
Generates a list of frequent words from the list of messages.
Uses a list of common words to avoid.
Takes one message at a time and repeats the following steps:
    Get individual words by splitting the message on every space.
    Convert each word to lower case.
    If the word is in the list of common words, go to the next word.
    Otherwise, "clean" the word. This involves removing any non-alphanumeric characters.
    If the word is already in the dictionary, increment the count.
    Else add it to the list.
When the entire list of messages is covered, return the frequency dict.
"""
def get_word_frequency(message_list):
    with open('commonWords.txt', 'r') as word_file:
        common_words_list = word_file.read()
    clean_word = re.compile(r'[а-яА-Яa-zA-z0-9]+')
    frequency_dictionary = Counter()
    for messages in message_list:
        processed_words = []
        words_list = messages.split(' ')
        for words in words_list:
            word = words.lower()
            if word in common_words_list:
                continue
            if clean_word.search(word):
                word = clean_word.search(word).group()
                processed_words.append(word)
        frequency_dictionary.update(processed_words)
    return frequency_dictionary

"""
Function to create a new sheet or find an existing sheet.
Returns a Worksheet object as well as the Workbook object.
First, it checks if the file: data.xlsx already exists.
If the file exists, just create a new sheet with the given sheet_name argument.
Otherwise, create a new Workbook and set the sheet name.
Return the Workbook, Worksheet
"""
def get_workbook_and_sheet(sheet_name, output_file):
    if os.path.isfile(output_file):
        xl = openpyxl.load_workbook(output_file)
        xl.create_sheet(title=sheet_name)
        sheet = xl[sheet_name]
    else:
        xl = openpyxl.Workbook()
        sheet = xl.active
        sheet.title = sheet_name
    return xl, sheet

"""
Function to save a dictionary to an Excel file.
Returns nothing.
Calls the get_workbook_and_sheet() function to get the workbook and sheet.
Sets the column width & headers before inserting each item of the dictionary to the sheet.
Then saves the workbook.
"""
def to_xl(dictionary, sheet_name, col1, col2, output_file):
    xl, sheet = get_workbook_and_sheet(sheet_name, output_file)
    # Column Width
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    # Column Headers
    sheet.cell(row=1, column=1).value = col1
    sheet.cell(row=2, column=2).value = col2
    # Insert Data
    for row, (key, value) in enumerate(dictionary.items()):
        sheet.cell(row=row + 2, column=1).value = key
        sheet.cell(row=row + 2, column=2).value = value
    xl.save(output_file)

"""
Function to accept command-line arguments.
If no arguments are found, then it prints an error and stops the script.
Else it returns the arguments.
"""
def get_file_name():
    #If no filename is given
    if len(sys.argv) < 2:
        print('Usage: analyze.py file_name.extension')
        sys.exit()
    #Get file name from command line
    return ' '.join(sys.argv[1:])

"""
Function to read the input file of chats.
Copies it to a variable and returns it.
"""
def read_file(file_name):
    with open(file_name, 'r', encoding='utf-8') as fi:
        text_to_analyze = fi.readlines()
    return text_to_analyze

"""
Converts the hour of the day from the 12-hour clock to the 24-hour clock.
"""
def to_24_hour_clock(hours, time_period):
    if time_period == ' PM' and int(hours) != 12:
        return int(hours) + 12
    elif time_period == ' AM' and int(hours) == 12:
        return '00'
    elif len(hours) == 1:
        return '0' + str(hours)
    else:
        return hours

"""
    Initialize all variables.
    Read one message at a time.
    Check if it matches our Regex "line" format.
    If it does, perform this process:
        Separate the date, time, person and put them into separate dictionaries.
        If the message isn't some sort of media attachment, add to message_list.
        Increment the number of messages.
    """
def collect_data(text_to_analyze):
    # Initializing our variables
    number_of_messages = 0
    date_dictionary = {}
    time_dictionary = {}
    person_dictionary = {}
    message_list = []

    # Collecting data into variables
    for lines in text_to_analyze:
        if message_format.search(lines):
            found = message_format.search(lines)
            date_dictionary = add_to_dictionary(date_dictionary, found[2])
            person_dictionary = add_to_dictionary(person_dictionary, found[10])
            #Time
            time = time_split.search(found[4])
            hours = to_24_hour_clock(time[1], found[7])
            time_dictionary = add_to_dictionary(time_dictionary, str(hours))
            #Message. Ignore media.
            if '<Media omitted>' not in found[12] and '<‎attached>' not in found[12]:
                message_list.append(found[12])
            number_of_messages += 1

    word_dictionary = get_word_frequency(message_list)
    return date_dictionary, time_dictionary, person_dictionary, word_dictionary, number_of_messages

"""
Sorts a dictionary by value or key. Value by default.
Uses an OrderedDict since in Python dictionaries can't be sorted.
"""
def sort_dictionary(dictionary, sort_by='value'):
    if sort_by == 'key':
        return OrderedDict(sorted(dictionary.items()))
    return OrderedDict(sorted(dictionary.items(), key=lambda x:x[1], reverse=True))

def driver():
    # Read given file
    file_name_with_extension = get_file_name()
    file_name = file_split.search(file_name_with_extension)[1]
    text_to_analyze = read_file(file_name_with_extension)

    # Collect data
    date_dictionary, time_dictionary, person_dictionary, word_dictionary, number_of_messages = collect_data(text_to_analyze)

    # Sort all Dictionaries here
    word_dictionary = OrderedDict(word_dictionary.most_common(20))
    date_dictionary = sort_dictionary(date_dictionary)
    time_dictionary = sort_dictionary(time_dictionary, 'key')
    person_dictionary = sort_dictionary(person_dictionary)

    if not os.path.exists('output'):
        os.mkdir('output')

    #Generate graphs
    graphs.histogram(
       time_dictionary,
       'Message Frequency Chart in ' + file_name,
       'output/' + file_name + '-time_activity.png'
    )
    graphs.bar_graph(
        word_dictionary, 20, 'Uses',
        'Most used words in ' + str(number_of_messages) + ' messages in ' + file_name,
        'output/' + file_name + '-word_frequency.png'
    )
    graphs.bar_graph(
        date_dictionary, 20, 'Messages',
        'Most Messages in ' + file_name,
        'output/' + file_name + '-date_activity.png'
    )
    graphs.bar_graph(
        person_dictionary, 20, 'Messages',
        'Most active person in ' + file_name,
        'output/' + file_name + '-person_activity.png'
    )

    # Remove old data sheets
    output_file = 'output/' + file_name + '-data.xlsx'
    if os.path.isfile(output_file):
        os.unlink(output_file)

    #Add to excel sheet
    to_xl(date_dictionary, 'Dates', 'Date', 'No. of Messages', output_file)
    to_xl(person_dictionary, 'People', 'Sender', 'No. of Messages', output_file)
    to_xl(time_dictionary, 'Times', 'Time', 'No. of Messages', output_file)
    to_xl(word_dictionary, 'Words', 'Word', 'No. of Occurences', output_file)

if __name__ == "__main__":
    driver()
